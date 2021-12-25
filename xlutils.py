import openpyxl


def getRowCount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return (sheet.max_row)


def getColumnCount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return (sheet.max_col)


def readdata(file, SheetName,rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return sheet.cell(row=rownum,column=columnno).value


def writedata(file, SheetName,rownum, columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    sheet.cell(row=rownum,column=columnno).value = data
    workbook.save(file)